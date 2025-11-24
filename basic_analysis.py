import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================= CONFIGURATION =================
DATA_FOLDER = "Datasets"
OUTPUT_FOLDER = "Output_Basic"
OUTPUT_FILE = "Basic_Statistics.xlsx"
TRIM_SECONDS = 1.0  

if not os.path.exists(OUTPUT_FOLDER):
    os.makedirs(OUTPUT_FOLDER)

# ================= HELPERS =================
def load_phyphox_xls(filepath):
    try:
        df_acc = pd.read_excel(filepath, sheet_name='Accelerometer')
        df_gyro = pd.read_excel(filepath, sheet_name='Gyroscope')
        
        # Standardize Time
        acc_t = [c for c in df_acc.columns if "Time" in c][0]
        gyro_t = [c for c in df_gyro.columns if "Time" in c][0]
        df_acc.rename(columns={acc_t: "Time"}, inplace=True)
        df_gyro.rename(columns={gyro_t: "Time"}, inplace=True)
        
        # Merge
        df_merged = df_acc.copy()
        for col in df_gyro.columns:
            if col != "Time":
                df_merged[col] = np.interp(df_acc["Time"], df_gyro["Time"], df_gyro[col])
        
        return df_merged
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return None

def process_and_plot(df, filename):
    # 1. Trim Data
    t_min = df["Time"].min() + TRIM_SECONDS
    t_max = df["Time"].max() - TRIM_SECONDS
    df_trim = df[(df["Time"] >= t_min) & (df["Time"] <= t_max)].copy()
    
    if df_trim.empty: return None

    # 2. Plot (Optional: Comment out to speed up)
    plt.figure(figsize=(10, 5))
    for axis in ['x', 'y', 'z']:
        col = [c for c in df_trim.columns if f"Acceleration {axis}" in c][0]
        plt.plot(df_trim["Time"], df_trim[col], label=f"Acc {axis}")
    plt.title(f"Raw Accelerometer - {filename}")
    plt.ylabel("Acc (m/s^2)")
    plt.xlabel("Time (s)")
    plt.legend()
    plt.grid(True)
    plt.savefig(os.path.join(OUTPUT_FOLDER, f"{filename}_Acc.png"))
    plt.close()

    plt.figure(figsize=(10, 5))
    for axis in ['x', 'y', 'z']:
        col = [c for c in df_trim.columns if f"Gyroscope {axis}" in c][0]
        plt.plot(df_trim["Time"], df_trim[col], label=f"Gyro {axis}")
    plt.title(f"Raw Gyroscope - {filename}")
    plt.ylabel("Rad/s")
    plt.xlabel("Time (s)")
    plt.legend()
    plt.grid(True)
    plt.savefig(os.path.join(OUTPUT_FOLDER, f"{filename}_Gyro.png"))
    plt.close()

    # 3. Calculate Stats
    stats = {"Filename": filename}
    for col in df_trim.columns:
        if col != "Time":
            stats[f"{col}_Mean"] = df_trim[col].mean()
            stats[f"{col}_Std"] = df_trim[col].std()
    return stats

def save_formatted_excel(stats_list, output_path):
    if not stats_list: return

    # 1. Create Base DataFrame
    df = pd.DataFrame(stats_list)
    
    # 2. Calculate Group Averages (bump1, bump2 -> bump_avg)
    # Regex extracts the alphabetic prefix (e.g., "bump" from "bump1")
    df['Group'] = df['Filename'].apply(lambda x: re.match(r"([a-zA-Z]+)", x).group(1) if re.match(r"([a-zA-Z]+)", x) else "Other")
    
    # Calculate Mean for each group
    group_means = df.groupby('Group').mean(numeric_only=True).reset_index()
    group_means['Filename'] = group_means['Group'] + "_avg" # Rename to "bump_avg"
    
    # Cleanup
    df = df.drop(columns=['Group'])
    group_means = group_means.drop(columns=['Group'])
    
    # Align columns
    group_means = group_means[df.columns]
    
    # 3. Combine: Raw Data TOP, Averages BOTTOM
    df_final = pd.concat([df, group_means], ignore_index=True)

    # --- WRITE WITH STYLING ---
    print(f"Saving formatted report to {output_path}...")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Statistics')
        
        workbook = writer.book
        worksheet = writer.sheets['Statistics']
        
        # Define Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid") # Dark Blue
        
        avg_row_font = Font(bold=True)
        avg_row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Light Yellow
        
        center_align = Alignment(horizontal='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply Header Style
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Apply Row Styles
        # Iterate through all rows (start=2 to skip header)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            # Check if this is an Average Row (contains "_avg")
            filename_cell = row[0] # Column A
            is_avg_row = "_avg" in str(filename_cell.value)
            
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align
                
                if isinstance(cell.value, float):
                    cell.number_format = '0.0000'

                if is_avg_row:
                    cell.font = avg_row_font
                    cell.fill = avg_row_fill

        # Auto-Adjust Widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max_length + 2

# ================= MAIN EXECUTION =================
files = [f for f in os.listdir(DATA_FOLDER) if f.endswith(('.xls', '.xlsx'))]
all_stats = []

print(f"Found {len(files)} files. Processing...")

for f in files:
    fname = os.path.splitext(f)[0]
    df = load_phyphox_xls(os.path.join(DATA_FOLDER, f))
    if df is not None:
        res = process_and_plot(df, fname)
        if res:
            all_stats.append(res)
            print(f"Processed: {fname}")

# Save Attractive Excel
save_formatted_excel(all_stats, os.path.join(OUTPUT_FOLDER, OUTPUT_FILE))
print(f"\nDone! Report saved as '{OUTPUT_FOLDER}/{OUTPUT_FILE}'.")